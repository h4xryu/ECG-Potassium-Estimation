import os
import csv
import numpy as np 
from openpyxl import load_workbook
from .manage_dirs import *
from numpy.typing import NDArray
from typing import Callable, Tuple, List, Any

def generate_SPL_datas(root="./dataset/SPL",xlsx="Hyperkalemia_level.xlsx"):
    '''
    환자번호 + 정상(A) or 비정상(B)
    리스트 파일을 txt로 저장

    환자번호와 A,B 에 대응하는 넘파이로 저장
    '''

    # Excel 파일 열기
    file_path = os.path.join(os.path.join(root,'SPL'),xlsx)
    workbook = load_workbook(file_path)

    # # 특정 시트 불러오기
    sheet = workbook[workbook.sheetnames[0]]

    # 데이터 읽기
    spl_list = []
    with open(file_path[:-5]+'.txt', 'w') as reset:
        reset.write('')
    skip = 0
    for num, spl_a, spl_b in sheet.iter_rows(values_only=True):
        if skip == 0:
            skip = 1
            continue
        if spl_a is not None:
            with open(file_path[:-5]+'.txt', 'a') as pat_txt:
                pat_txt.write(str(num)+'A\n')
            spl_list.append(spl_a)

        if spl_b is not None:
            with open(file_path[:-5]+'.txt', 'a') as pat_txt:
                pat_txt.write(str(num)+'B\n')
            spl_list.append(spl_b)
    
    np.save(file_path[:-5]+'.npy',np.array(spl_list))

@DirectoryProcess
def load_xlsx_datas(root="./dataset/") -> (NDArray[np.float64], str):
    '''
    :param root: 데이터 셋이 있는 디렉토리 입력
    :return: SPL, 환자번호 리턴
    '''

    npy_files = sorted([f for f in os.listdir(root) if f.endswith(".npy")])
    txt_files = sorted([f for f in os.listdir(root) if f.endswith(".txt")])

    npy_path = os.path.join(root, npy_files[0])
    txt_path = os.path.join(root, txt_files[0])

    with open(txt_path, 'r') as txt:
            lines = txt.readlines()
    lines = [x.strip() for x in lines]
    arr = np.load(npy_path)

    #환자와 혈청칼륨수치
    # for SPL, patient in zip(arr, lines):
    #     print(SPL)
    #     print(patient)

        
    return arr, lines
        

